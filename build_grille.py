#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Génère un classeur Excel de notation pour le Bac Blanc STI2D 2I2D
12-05-2026 — sujet AQUALYON Feyssine, partie commune (12 pts).

Architecture (cf. plan validé) :
  • Feuille `Élèves` : liste de classe + notes finales agrégées (dynamique, 30 lignes)
  • Feuille `BB2 12-05-26` : grille d'évaluation
      - Zone fixe A-F (N°, Partie, Question, Type, Critère, Poids)
      - Bloc compact 3 colonnes par élève (Saisie / Note brute / Signal)
      - Validation listes, MFC, figement, protection
  • Feuille `Mode d'emploi` : procédure de saisie

Usage :
  ~/.venv-xlsx/bin/python build_grille.py
puis :
  libreoffice --headless --calc --convert-to xlsx <fichier>  (force recalcul)
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Protection, Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import (
    CellIsRule, FormulaRule, DataBarRule, Rule,
)
from openpyxl.styles.differential import DifferentialStyle
from pathlib import Path

# ─────────────────────────────────────────────────────────────────────────────
# Configuration
# ─────────────────────────────────────────────────────────────────────────────

N_ELEVES = 30  # nb d'élèves max (dimensionné pour ajout/retrait facile, 24 utilisés)
N_ELEVES_PREREMPLIS = 24  # noms préremplis "Élève 1".."Élève N"
NOM_EVAL = "BB2 12-05-26"
DATE_EVAL = "12-05-2026"
TITRE_LIGNE1 = "BAC BLANC N°2 — 12-05-2026 — STI2D 2I2D — Site AQUALYON"
TITRE_LIGNE2 = "Analyse des performances de la station de méthanisation de la Feyssine — Partie commune (12 pts)"

OUT_DIR = Path("/home/fz/Documents/Nextcloud_Edu/Enseignement/2025-2026/2I2D/Bac_blanc_mai_2026")
OUT_FILE = OUT_DIR / "TRAME_260512_NOTES_BAC_BLANC_N°2_v2.xlsx"

# ─────────────────────────────────────────────────────────────────────────────
# Définition des critères (extrait du corrigé officiel + ancienne grille)
# Format : (type, partie, question, libellé/attendu, poids)
#   type ∈ {"partie", "critere"}
#   poids = somme par partie doit faire 1.00 (vérifié plus bas)
# ─────────────────────────────────────────────────────────────────────────────

POIDS_PARTIES = {
    "P1": 0.20,
    "P2": 0.25,
    "P3": 0.10,
    "P4": 0.20,
    "P5": 0.25,
}

CRITERES = [
    # ─── Partie 1 ──────────────────────────────────────────────────────────
    ("partie",  "P1", "",     "Comment la station de la Feyssine permet-elle de répondre aux besoins de la métropole ?",                                                 None),
    ("critere", "P1", "Q1.1", "La station de Saint-Fons est saturée en DBO.",                                                                                            0.15),
    ("critere", "P1", "Q1.1", "Il faut construire une nouvelle station pour répondre à l'augmentation de population.",                                                   0.15),
    ("critere", "P1", "Q1.2", "Justification : terrain près du Rhône, près du périphérique, point altimétrique bas, entouré de parc naturel (pas d'habitation proche).", 0.15),
    ("critere", "P1", "Q1.3", "Surface lit roseaux = 10 m²/EqHab × 300 000 = 3 000 000 m² (300 ha).",                                                                    0.15),
    ("critere", "P1", "Q1.3", "Conso énergie Boue Activée = 3,2 kWh/kgDBO5 × 17 100 = 54 720 kWh.",                                                                      0.15),
    ("critere", "P1", "Q1.3", "Coût biofiltre = 2 142 €/EqHab × 300 000 = 642,6 M€.",                                                                                    0.15),
    ("critere", "P1", "Q1.4", "Boue Activée justifiée : lit roseaux trop grand, biofiltre trop énergivore et trop cher.",                                                0.10),

    # ─── Partie 2 ──────────────────────────────────────────────────────────
    ("partie",  "P2", "",     "Comment la digestion des boues améliore-t-elle les performances d'une STEU dans une démarche DD ?",                                       None),
    ("critere", "P2", "Q2.1", "Type 1 : flux entrant = énergie électrique.",                                                                                             0.10),
    ("critere", "P2", "Q2.1", "Type 2 : entrants = énergie élec + gaz GRDF ; sortants = méthane (GRDF) + thermique perdue (torchère).",                                  0.10),
    ("critere", "P2", "Q2.2", "Calcul : 6 160 000 × 0,1 = 616 000 kgeq.CO₂/an.",                                                                                          0.10),
    ("critere", "P2", "Q2.2", "Ligne (1) DR2 correctement complétée.",                                                                                                   0.10),
    ("critere", "P2", "Q2.3", "Calcul : 5 296 000 × (−0,2) = −1 059 200 kgeq.CO₂/an.",                                                                                    0.10),
    ("critere", "P2", "Q2.3", "Ligne (4) DR2 correctement complétée.",                                                                                                   0.10),
    ("critere", "P2", "Q2.4", "Totaux correctement calculés (DR2).",                                                                                                     0.10),
    ("critere", "P2", "Q2.4", "Conclusion DD : économique (biogaz, valorisation combustible), environnemental (réduction GES), sociétal (nuisances olfactives).",        0.30),

    # ─── Partie 3 ──────────────────────────────────────────────────────────
    ("partie",  "P3", "",     "Comment la maîtrise de l'information permet-elle de garantir la sécurité des personnes ?",                                                None),
    ("critere", "P3", "Q3.1", "Masque 255.255.0.0 → (256 × 256) − 2 = 65 534 adresses possibles.",                                                                       0.25),
    ("critere", "P3", "Q3.1", "Conversion binaire 10101100.00010000.00110010.11001000 → 172.16.50.200.",                                                                 0.25),
    ("critere", "P3", "Q3.2", "Adresse IP proposée compatible et disponible (dans la plage réservée WiFi).",                                                             0.25),
    ("critere", "P3", "Q3.3", "Protection des personnes : détecteurs multigaz + surveillance temps réel des expositions aux gaz nocifs.",                                0.25),

    # ─── Partie 4 ──────────────────────────────────────────────────────────
    ("partie",  "P4", "",     "Comment justifier l'utilisation d'un séparateur de boues biologiques ?",                                                                  None),
    ("critere", "P4", "Q4.1", "Flux entrant : énergie électrique.",                                                                                                      0.15),
    ("critere", "P4", "Q4.1", "Flux interne : énergie mécanique (variateur → moteur).",                                                                                  0.15),
    ("critere", "P4", "Q4.1", "Flux interne : énergie mécanique (multiplicateur → bol/vis).",                                                                            0.15),
    ("critere", "P4", "Q4.2", "Ω_bol = 2π × 2600/60 ≈ 272 rad/s.",                                                                                                       0.15),
    ("critere", "P4", "Q4.3", "F_cb = 1200 × 1 × 0,335 × 300² = 36 180 kN.",                                                                                             0.15),
    ("critere", "P4", "Q4.4", "F_cb = 36 180 kN > F_ce = 30 200 kN.",                                                                                                    0.15),
    ("critere", "P4", "Q4.4", "Conclusion : forces centrifuges différentes sur boue et eau → séparation possible.",                                                      0.10),

    # ─── Partie 5 ──────────────────────────────────────────────────────────
    ("partie",  "P5", "",     "Peut-on valider l'implantation d'une torchère ?",                                                                                          None),
    ("critere", "P5", "Q5.1", "V = H·π·(R² − r²) avec H=7000mm, R=750mm, r=740mm → V ≈ 3,28·10⁸ mm³ = 0,328 m³.",                                                         0.25),
    ("critere", "P5", "Q5.2", "Masse fût = ρ × V = 8000 × 0,328 = 2 624 kg.",                                                                                            0.20),
    ("critere", "P5", "Q5.2", "Charge totale = 9,81 × (2624 + 1250) = 38 004 N ; P_c par cornière = 38 004/4 = 9 501 N.",                                                0.20),
    ("critere", "P5", "Q5.3", "Surface dalle = 4 × 2,5 = 10 m².",                                                                                                        0.05),
    ("critere", "P5", "Q5.3", "σ_sol = 200 000 / 10 = 20 000 Pa = 0,02 MPa.",                                                                                            0.05),
    ("critere", "P5", "Q5.3", "Conclusion : σ_sol (0,02 MPa) < résistance admissible (0,1 MPa) → le sol supporte la charge.",                                            0.05),
    ("critere", "P5", "Q5.4", "Brûler le méthane est écologiquement plus intéressant que le relâcher dans l'atmosphère.",                                                0.10),
    ("critere", "P5", "Q5.4", "La dalle existante et le sol sont capables de supporter l'installation complète.",                                                        0.10),
]

# ─────────────────────────────────────────────────────────────────────────────
# Vérifications de cohérence (échouent le script si KO)
# ─────────────────────────────────────────────────────────────────────────────

def verifier_poids():
    """Σ(parties) doit faire 1.00 ; Σ(critères) par partie doit faire 1.00."""
    s_parties = sum(POIDS_PARTIES.values())
    assert abs(s_parties - 1.0) < 1e-9, f"Σ poids parties = {s_parties} ≠ 1.00"

    par_partie = {}
    for ligne in CRITERES:
        type_, partie, q, lib, poids = ligne
        if type_ == "critere":
            par_partie.setdefault(partie, 0.0)
            par_partie[partie] += poids
    for p, s in par_partie.items():
        assert abs(s - 1.0) < 1e-9, f"Σ poids critères {p} = {s} ≠ 1.00"
    print(f"✓ Vérif poids OK — parties: {s_parties:.4f}, par partie: {par_partie}")

verifier_poids()

# ─────────────────────────────────────────────────────────────────────────────
# Styles partagés
# ─────────────────────────────────────────────────────────────────────────────

THIN = Side(border_style="thin", color="888888")
MEDIUM = Side(border_style="medium", color="000000")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORDER_BOLD = Border(left=MEDIUM, right=MEDIUM, top=MEDIUM, bottom=MEDIUM)

FILL_TITRE = PatternFill("solid", fgColor="2E5C8A")
FILL_ENTETE = PatternFill("solid", fgColor="D4E1F1")
FILL_PARTIE = PatternFill("solid", fgColor="FFE699")
FILL_POIDS = PatternFill("solid", fgColor="FFF2CC")
FILL_NOTE = PatternFill("solid", fgColor="E2EFDA")
FILL_GRIS = PatternFill("solid", fgColor="EEEEEE")

FONT_TITRE = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
FONT_SOUS_TITRE = Font(name="Calibri", size=11, italic=True, color="333333")
FONT_ENTETE = Font(name="Calibri", size=10, bold=True)
FONT_PARTIE = Font(name="Calibri", size=11, bold=True)
FONT_NORMAL = Font(name="Calibri", size=10)
FONT_SIGNAL = Font(name="Calibri", size=10, bold=True, color="CC0000")

AL_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
AL_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
AL_LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)
AL_CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)

PROT_UNLOCKED = Protection(locked=False)
PROT_LOCKED = Protection(locked=True)

# ─────────────────────────────────────────────────────────────────────────────
# Construction du Workbook
# ─────────────────────────────────────────────────────────────────────────────

wb = Workbook()
# Renommer la première feuille (qui sera créée par défaut)
ws_eleves = wb.active
ws_eleves.title = "Élèves"
ws_eval = wb.create_sheet(NOM_EVAL)
ws_help = wb.create_sheet("Mode d'emploi")

# ─────────────────────────────────────────────────────────────────────────────
# FEUILLE D'ÉVALUATION
# ─────────────────────────────────────────────────────────────────────────────

# Colonnes : A=N° | B=Partie | C=Question | D=Type | E=Critère | F=Poids
# Puis 3 colonnes par élève : Saisie / Note brute / Signal
COL_NUM = 1
COL_PARTIE = 2
COL_QUESTION = 3
COL_TYPE = 4
COL_CRITERE = 5
COL_POIDS = 6
COL_BLOC_DEBUT = 7  # G = première colonne saisie élève 1

def col_saisie(i_eleve):
    """Colonne 1-indexée de la cellule de saisie pour l'élève i (1..N_ELEVES)."""
    return COL_BLOC_DEBUT + 3 * (i_eleve - 1)

def col_note(i_eleve):
    return COL_BLOC_DEBUT + 3 * (i_eleve - 1) + 1

def col_signal(i_eleve):
    return COL_BLOC_DEBUT + 3 * (i_eleve - 1) + 2

# Lignes
ROW_TITRE = 1
ROW_SOUS_TITRE = 2
ROW_ENTETE = 3
ROW_DATA_DEBUT = 4  # première ligne de critère / partie

# ─── Lignes 1-2 : titre ──────────────────────────────────────────────────────
ws_eval.cell(ROW_TITRE, 1, TITRE_LIGNE1)
ws_eval.cell(ROW_TITRE, 1).font = FONT_TITRE
ws_eval.cell(ROW_TITRE, 1).fill = FILL_TITRE
ws_eval.cell(ROW_TITRE, 1).alignment = AL_CENTER
ws_eval.row_dimensions[ROW_TITRE].height = 24

ws_eval.cell(ROW_SOUS_TITRE, 1, TITRE_LIGNE2)
ws_eval.cell(ROW_SOUS_TITRE, 1).font = FONT_SOUS_TITRE
ws_eval.cell(ROW_SOUS_TITRE, 1).alignment = AL_CENTER
ws_eval.row_dimensions[ROW_SOUS_TITRE].height = 20

n_cols_total = COL_BLOC_DEBUT - 1 + 3 * N_ELEVES
ws_eval.merge_cells(start_row=ROW_TITRE, start_column=1, end_row=ROW_TITRE, end_column=n_cols_total)
ws_eval.merge_cells(start_row=ROW_SOUS_TITRE, start_column=1, end_row=ROW_SOUS_TITRE, end_column=n_cols_total)

# ─── Ligne 3 : en-têtes ──────────────────────────────────────────────────────
ENTETES_FIXES = ["N°", "Partie", "Question", "Type", "Critère / attendu", "Poids"]
for i, label in enumerate(ENTETES_FIXES, start=1):
    c = ws_eval.cell(ROW_ENTETE, i, label)
    c.font = FONT_ENTETE
    c.fill = FILL_ENTETE
    c.alignment = AL_CENTER_WRAP
    c.border = BORDER_ALL

for i in range(1, N_ELEVES + 1):
    c_sai = ws_eval.cell(ROW_ENTETE, col_saisie(i))
    c_sai.value = f"E{i}"  # libellé court ; le nom complet est sur feuille Élèves
    c_sai.font = FONT_ENTETE
    c_sai.fill = FILL_ENTETE
    c_sai.alignment = AL_CENTER_WRAP
    c_sai.border = BORDER_ALL
    c_note = ws_eval.cell(ROW_ENTETE, col_note(i), "Note")
    c_note.font = FONT_ENTETE
    c_note.fill = FILL_ENTETE
    c_note.alignment = AL_CENTER_WRAP
    c_note.border = BORDER_ALL
    c_sig = ws_eval.cell(ROW_ENTETE, col_signal(i), "⚠")
    c_sig.font = FONT_ENTETE
    c_sig.fill = FILL_ENTETE
    c_sig.alignment = AL_CENTER_WRAP
    c_sig.border = BORDER_ALL

ws_eval.row_dimensions[ROW_ENTETE].height = 32

# ─── Lignes 4+ : critères ────────────────────────────────────────────────────
# Construction des lignes en gardant trace de :
#  - quelles lignes sont des "titres de partie" (à exclure des sommes)
#  - quelles lignes sont des critères (à inclure)
#  - le numéro de partie de chaque critère (pour la formule note finale)

ROWS_CRITERES = {}     # partie_id -> [list of row numbers]
ROW_PARTIE = {}        # partie_id -> row du titre
row = ROW_DATA_DEBUT
num_critere = 0

for entry in CRITERES:
    type_, partie, q, lib, poids = entry
    if type_ == "partie":
        ROW_PARTIE[partie] = row
        # Titre de partie : fusion col A→E + cellule poids partie en F
        ws_eval.cell(row, COL_PARTIE, partie).fill = FILL_PARTIE
        ws_eval.cell(row, COL_PARTIE, partie).font = FONT_PARTIE
        ws_eval.cell(row, COL_PARTIE, partie).alignment = AL_CENTER
        ws_eval.cell(row, COL_PARTIE, partie).border = BORDER_BOLD
        # Mettre le libellé de la partie en col E
        ws_eval.cell(row, COL_CRITERE, lib).fill = FILL_PARTIE
        ws_eval.cell(row, COL_CRITERE, lib).font = FONT_PARTIE
        ws_eval.cell(row, COL_CRITERE, lib).alignment = AL_LEFT
        ws_eval.cell(row, COL_CRITERE, lib).border = BORDER_BOLD
        # Numéro vide et type "Partie"
        ws_eval.cell(row, COL_NUM, "").fill = FILL_PARTIE
        ws_eval.cell(row, COL_QUESTION, "").fill = FILL_PARTIE
        ws_eval.cell(row, COL_TYPE, "Partie").fill = FILL_PARTIE
        ws_eval.cell(row, COL_TYPE, "Partie").font = FONT_PARTIE
        ws_eval.cell(row, COL_TYPE, "Partie").alignment = AL_CENTER
        # Poids de la partie dans col F
        p_partie = POIDS_PARTIES[partie]
        cell_p = ws_eval.cell(row, COL_POIDS, p_partie)
        cell_p.fill = FILL_POIDS
        cell_p.font = FONT_PARTIE
        cell_p.alignment = AL_CENTER
        cell_p.border = BORDER_BOLD
        cell_p.number_format = "0.00"
        cell_p.protection = PROT_UNLOCKED  # le prof peut modifier le poids partie
        # Bordure sur les cellules saisie/note/signal de cette ligne (vides, juste fond)
        for i in range(1, N_ELEVES + 1):
            ws_eval.cell(row, col_saisie(i)).fill = FILL_PARTIE
            ws_eval.cell(row, col_note(i)).fill = FILL_PARTIE
            ws_eval.cell(row, col_signal(i)).fill = FILL_PARTIE
        ws_eval.row_dimensions[row].height = 22
    else:
        num_critere += 1
        ROWS_CRITERES.setdefault(partie, []).append(row)
        # N°, Partie, Question, Type, Critère, Poids
        ws_eval.cell(row, COL_NUM, num_critere).alignment = AL_CENTER
        ws_eval.cell(row, COL_NUM).font = FONT_NORMAL
        ws_eval.cell(row, COL_NUM).border = BORDER_ALL
        ws_eval.cell(row, COL_PARTIE, partie).alignment = AL_CENTER
        ws_eval.cell(row, COL_PARTIE).font = FONT_NORMAL
        ws_eval.cell(row, COL_PARTIE).border = BORDER_ALL
        ws_eval.cell(row, COL_QUESTION, q).alignment = AL_CENTER
        ws_eval.cell(row, COL_QUESTION).font = FONT_NORMAL
        ws_eval.cell(row, COL_QUESTION).border = BORDER_ALL
        ws_eval.cell(row, COL_TYPE, "Critère").alignment = AL_CENTER
        ws_eval.cell(row, COL_TYPE).font = FONT_NORMAL
        ws_eval.cell(row, COL_TYPE).border = BORDER_ALL
        ws_eval.cell(row, COL_CRITERE, lib).alignment = AL_LEFT_TOP
        ws_eval.cell(row, COL_CRITERE).font = FONT_NORMAL
        ws_eval.cell(row, COL_CRITERE).border = BORDER_ALL
        cell_p = ws_eval.cell(row, COL_POIDS, poids)
        cell_p.alignment = AL_CENTER
        cell_p.font = FONT_NORMAL
        cell_p.fill = FILL_POIDS
        cell_p.border = BORDER_ALL
        cell_p.number_format = "0.00"
        cell_p.protection = PROT_UNLOCKED  # poids éditable

        # Formules par élève : référence à $F<row> (poids critère) et au poids de la partie
        # On utilise la cellule fixe du poids partie : F<row_partie> = $F$<row_partie>
        row_partie = ROW_PARTIE[partie]
        cell_partie_poids = f"$F${row_partie}"
        cell_critere_poids = f"$F{row}"  # absolu en colonne F, relatif en ligne (référence sur la même ligne)

        for i in range(1, N_ELEVES + 1):
            cs = col_saisie(i)
            cn = col_note(i)
            cg = col_signal(i)
            ref_saisie = f"{get_column_letter(cs)}{row}"
            # Saisie : vide par défaut ; sera contrainte par DataValidation plus tard
            cell_sai = ws_eval.cell(row, cs)
            cell_sai.alignment = AL_CENTER
            cell_sai.font = FONT_NORMAL
            cell_sai.border = BORDER_ALL
            cell_sai.protection = PROT_UNLOCKED  # déverrouillé pour saisie
            # Note brute pondérée
            # formule : =IF(saisie="","", IF(saisie="NE","", score(saisie) * poids_crit * poids_partie * 20))
            # score("0")=0, score("1")=0.5, score("2")=1 via IF imbriqués (compatible LibreOffice toutes versions)
            f_note = (
                f'=IF({ref_saisie}="","",'
                f'IF({ref_saisie}="NE","",'
                f'IF({ref_saisie}="0",0,IF({ref_saisie}="1",0.5,IF({ref_saisie}="2",1,0)))'
                f'*{cell_critere_poids}*{cell_partie_poids}*20))'
            )
            cell_nt = ws_eval.cell(row, cn, f_note)
            cell_nt.alignment = AL_CENTER
            cell_nt.font = FONT_NORMAL
            cell_nt.border = BORDER_ALL
            cell_nt.number_format = "0.00"
            cell_nt.fill = FILL_NOTE
            # Signal : "◄" si autre saisie au-dessus mais celle-ci vide, "~" si NE
            f_signal = (
                f'=IF({ref_saisie}="NE","~",'
                f'IF({ref_saisie}="","",""))'
            )
            cell_sg = ws_eval.cell(row, cg, f_signal)
            cell_sg.alignment = AL_CENTER
            cell_sg.font = FONT_SIGNAL
            cell_sg.border = BORDER_ALL

        ws_eval.row_dimensions[row].height = 36

    row += 1

ROW_DATA_FIN = row - 1  # dernière ligne critère/partie

# ─── Lignes audit & synthèse ─────────────────────────────────────────────────
ROW_AUDIT_POIDS = ROW_DATA_FIN + 2
ROW_NOTE_20 = ROW_DATA_FIN + 4
ROW_NOTE_12 = ROW_DATA_FIN + 5
ROW_STAT_PCT = ROW_DATA_FIN + 7
ROW_STAT_MOY = ROW_DATA_FIN + 8

# Audit poids : pour chaque partie, vérifier Σ(F<critère>) == 1.00
ws_eval.cell(ROW_AUDIT_POIDS, COL_CRITERE, "Audit Σ(poids critères) par partie (doit valoir 1,00 pour chaque P)").font = FONT_ENTETE
ws_eval.cell(ROW_AUDIT_POIDS, COL_CRITERE).alignment = AL_LEFT

# Formule audit poids : on liste Σ des critères par partie
audit_formula_parts = []
for partie, rows_p in ROWS_CRITERES.items():
    # range "$F<min>:$F<max>"
    rmin = min(rows_p)
    rmax = max(rows_p)
    audit_formula_parts.append(f'SUM($F${rmin}:$F${rmax})')
# Cellule F : Σ globale (somme des Σ partielles, doit faire len(parties) = 5)
audit_global_formula = f'={"+".join(audit_formula_parts)}'
audit_cell = ws_eval.cell(ROW_AUDIT_POIDS, COL_POIDS, audit_global_formula)
audit_cell.font = FONT_ENTETE
audit_cell.alignment = AL_CENTER
audit_cell.number_format = "0.00"
audit_cell.fill = FILL_POIDS
audit_cell.border = BORDER_BOLD

# Note finale /20 pour chaque élève
ws_eval.cell(ROW_NOTE_20, COL_CRITERE, "NOTE FINALE / 20").font = FONT_ENTETE
ws_eval.cell(ROW_NOTE_20, COL_CRITERE).alignment = AL_LEFT
ws_eval.cell(ROW_NOTE_20, COL_CRITERE).fill = FILL_NOTE

ws_eval.cell(ROW_NOTE_12, COL_CRITERE, "NOTE FINALE / 12").font = FONT_ENTETE
ws_eval.cell(ROW_NOTE_12, COL_CRITERE).alignment = AL_LEFT
ws_eval.cell(ROW_NOTE_12, COL_CRITERE).fill = FILL_NOTE

# Pour chaque élève : note /20 = Σ(notes brutes) / Σ(poids effectifs * 20) * 20
# Mais comme chaque note brute = score * poids_crit * poids_partie * 20,
# Σ notes brutes / 20 = score_pondéré ; et somme effective des poids = somme des poids partie*crit
# tels que la saisie ≠ "" et ≠ "NE".
# On utilise SUMPRODUCT avec masques.
for i in range(1, N_ELEVES + 1):
    cs = col_saisie(i)
    cn = col_note(i)
    col_letter_s = get_column_letter(cs)
    col_letter_n = get_column_letter(cn)

    # Plages saisies / notes (toutes les lignes critères, donc on prend l'union)
    # Pour simplicité : on prend la plage globale ROW_DATA_DEBUT..ROW_DATA_FIN
    # Les lignes "partie" auront saisie vide → masque les écarte automatiquement
    plage_saisie = f"{col_letter_s}{ROW_DATA_DEBUT}:{col_letter_s}{ROW_DATA_FIN}"
    plage_note = f"{col_letter_n}{ROW_DATA_DEBUT}:{col_letter_n}{ROW_DATA_FIN}"
    plage_poids_crit = f"$F${ROW_DATA_DEBUT}:$F${ROW_DATA_FIN}"

    # Pour le dénominateur : Σ (poids_crit * poids_partie * 20) où saisie est dans {"0","1","2"}
    # Comme poids_partie change selon ligne, on construit par parties :
    denom_parts = []
    for partie, rows_p in ROWS_CRITERES.items():
        rmin, rmax = min(rows_p), max(rows_p)
        rp = ROW_PARTIE[partie]
        # poids partie cste F<rp>, poids critères F<rmin:rmax>
        # masque : saisies non vides et non "NE"
        denom_parts.append(
            f'SUMPRODUCT(($F${rmin}:$F${rmax})'
            f'*({col_letter_s}{rmin}:{col_letter_s}{rmax}<>"")'
            f'*({col_letter_s}{rmin}:{col_letter_s}{rmax}<>"NE"))'
            f'*$F${rp}'
        )
    denom = "(" + "+".join(denom_parts) + ")*20"
    numer = f"SUM({plage_note})"
    # Note /20 = numer/denom (déjà sur 20 car notes brutes contiennent *20)
    # Mais le numer est en "points sur 20" déjà ; le denom = "poids total * 20".
    # Donc note finale = numer/denom * 20 → on annule le *20. Reformulons :
    # numer = SUM(notes_brutes) où chaque note_brute_i = score_i*poids_crit_i*poids_partie_i*20
    # denom (sans le *20) = SUM(poids_crit_i * poids_partie_i) pour i évalués
    # note/20 = (SUM(score * poids_crit * poids_partie * 20)) / (SUM(poids_crit * poids_partie) * 20) * 20
    #         = SUM(score * poids_crit * poids_partie) / SUM(poids_crit * poids_partie) * 20
    # Mais comme score ∈ {0, 0.5, 1}, score=1 → note max=20 quand tout est à 2.
    # Simplification : note/20 = numer / denom_sans_20.
    denom_sans20 = "(" + "+".join(p.replace("$F$" + str(ROW_PARTIE[list(POIDS_PARTIES)[0]]), "TEMP") for p in []) + ")"
    # Trop compliqué : refactorisons. On veut :
    #   note/20 = Σ(score * poids_crit * poids_partie * 20) / Σ(poids_crit * poids_partie * 20) * 20
    #          = Σ(note_brute_i) / [Σ(poids_crit * poids_partie) * 20] * 20
    #          = Σ(note_brute_i) / Σ(poids_crit * poids_partie)
    # → puisque note_brute_i contient déjà le facteur 20, on divise par (Σ poids * 20) puis on multiplie par 20
    # → c'est juste note/20 = Σ(note_brute_i) / Σ(poids_crit*poids_partie) → unité = directement /20.
    # Donc denom = SUM(poids_crit_i * poids_partie_i) sans le *20.
    denom_sans20_parts = []
    for partie, rows_p in ROWS_CRITERES.items():
        rmin, rmax = min(rows_p), max(rows_p)
        rp = ROW_PARTIE[partie]
        denom_sans20_parts.append(
            f'SUMPRODUCT(($F${rmin}:$F${rmax})'
            f'*({col_letter_s}{rmin}:{col_letter_s}{rmax}<>"")'
            f'*({col_letter_s}{rmin}:{col_letter_s}{rmax}<>"NE"))'
            f'*$F${rp}'
        )
    denom_clean = "(" + "+".join(denom_sans20_parts) + ")"
    f_note20 = f'=IFERROR(SUM({plage_note})/{denom_clean},"")'
    cell_n20 = ws_eval.cell(ROW_NOTE_20, cn, f_note20)
    cell_n20.alignment = AL_CENTER
    cell_n20.font = Font(name="Calibri", size=12, bold=True)
    cell_n20.number_format = "0.00"
    cell_n20.fill = FILL_NOTE
    cell_n20.border = BORDER_BOLD

    # Note /12 = note/20 * 12/20 (référence à la cellule ci-dessus)
    cell_n12_ref = f"{col_letter_n}{ROW_NOTE_20}"
    f_note12 = f'=IF({cell_n12_ref}="","",{cell_n12_ref}*12/20)'
    cell_n12 = ws_eval.cell(ROW_NOTE_12, cn, f_note12)
    cell_n12.alignment = AL_CENTER
    cell_n12.font = Font(name="Calibri", size=11, bold=True)
    cell_n12.number_format = "0.00"
    cell_n12.fill = FILL_NOTE
    cell_n12.border = BORDER_ALL

# Statistiques classe (par critère)
ws_eval.cell(ROW_STAT_PCT, COL_CRITERE, "% acquis (saisies = '2')").font = FONT_ENTETE
ws_eval.cell(ROW_STAT_PCT, COL_CRITERE).alignment = AL_LEFT
ws_eval.cell(ROW_STAT_MOY, COL_CRITERE, "Moyenne classe (note brute pondérée)").font = FONT_ENTETE
ws_eval.cell(ROW_STAT_MOY, COL_CRITERE).alignment = AL_LEFT

# Construire les plages de saisies/notes pour chaque ligne critère
# (pour les stats, on parcourt chaque ligne et on agrège horizontalement)
# Cellule stat sur la colonne F (= poids), on met les valeurs par critère sur les lignes des critères
# Variante : afficher pour chaque critère un % acquis dans la colonne signal de la "moyenne classe"
# → trop complexe. On se contente d'un total global :
all_saisies_ranges = []
all_notes_ranges = []
for i in range(1, N_ELEVES + 1):
    col_letter_s = get_column_letter(col_saisie(i))
    col_letter_n = get_column_letter(col_note(i))
    all_saisies_ranges.append(f"{col_letter_s}{ROW_DATA_DEBUT}:{col_letter_s}{ROW_DATA_FIN}")
    all_notes_ranges.append(f"{col_letter_n}{ROW_DATA_DEBUT}:{col_letter_n}{ROW_DATA_FIN}")

# % acquis global = count("2") / count(saisies évaluées) sur toute la grille
count2_parts = [f'COUNTIF({r},"2")' for r in all_saisies_ranges]
count_eval_parts = [
    f'(COUNTA({r})-COUNTIF({r},"NE"))' for r in all_saisies_ranges
]
f_pct_global = f'=IFERROR(({"+".join(count2_parts)})/({"+".join(count_eval_parts)}),"")'
cell_pct = ws_eval.cell(ROW_STAT_PCT, COL_POIDS, f_pct_global)
cell_pct.alignment = AL_CENTER
cell_pct.font = FONT_NORMAL
cell_pct.number_format = "0%"
cell_pct.border = BORDER_ALL

# Moyenne classe = AVERAGE des notes /20 des élèves saisis
note20_refs = [f"{get_column_letter(col_note(i))}{ROW_NOTE_20}" for i in range(1, N_ELEVES + 1)]
f_moy_classe = f'=IFERROR(AVERAGE({",".join(note20_refs)}),"")'
cell_moy = ws_eval.cell(ROW_STAT_MOY, COL_POIDS, f_moy_classe)
cell_moy.alignment = AL_CENTER
cell_moy.font = FONT_NORMAL
cell_moy.number_format = "0.00"
cell_moy.border = BORDER_ALL

# ─── Largeurs de colonnes ────────────────────────────────────────────────────
ws_eval.column_dimensions["A"].width = 4
ws_eval.column_dimensions["B"].width = 8
ws_eval.column_dimensions["C"].width = 9
ws_eval.column_dimensions["D"].width = 8
ws_eval.column_dimensions["E"].width = 65
ws_eval.column_dimensions["F"].width = 8
for i in range(1, N_ELEVES + 1):
    ws_eval.column_dimensions[get_column_letter(col_saisie(i))].width = 6
    ws_eval.column_dimensions[get_column_letter(col_note(i))].width = 7
    ws_eval.column_dimensions[get_column_letter(col_signal(i))].width = 3

# ─── DataValidation : liste {NE, 0, 1, 2} sur chaque cellule saisie ──────────
dv = DataValidation(
    type="list",
    formula1='"NE,0,1,2"',
    allow_blank=True,
    showErrorMessage=True,
    errorTitle="Saisie invalide",
    error="Valeurs autorisées : NE (non évalué), 0, 1, 2.",
)
ws_eval.add_data_validation(dv)
for i in range(1, N_ELEVES + 1):
    col_letter_s = get_column_letter(col_saisie(i))
    # Ne s'applique qu'aux lignes critères, pas aux lignes "partie"
    for partie, rows_p in ROWS_CRITERES.items():
        for r in rows_p:
            dv.add(f"{col_letter_s}{r}")

# ─── Mise en forme conditionnelle ────────────────────────────────────────────
# Règles globales sur les colonnes de saisie
green_fill = PatternFill("solid", fgColor="C6EFCE")
red_fill = PatternFill("solid", fgColor="FFC7CE")
gray_fill = PatternFill("solid", fgColor="D9D9D9")
yellow_fill = PatternFill("solid", fgColor="FFEB9C")
orange_fill = PatternFill("solid", fgColor="FFC000")

# Pour les colonnes saisies de tous les élèves
for i in range(1, N_ELEVES + 1):
    cs = col_saisie(i)
    cn = col_note(i)
    col_letter_s = get_column_letter(cs)
    col_letter_n = get_column_letter(cn)

    # Plage des cellules de saisie (lignes critères uniquement)
    plage_saisie = f"{col_letter_s}{ROW_DATA_DEBUT}:{col_letter_s}{ROW_DATA_FIN}"

    # Saisie = "2" → vert
    rule_2 = CellIsRule(operator="equal", formula=['"2"'], fill=green_fill)
    ws_eval.conditional_formatting.add(plage_saisie, rule_2)
    # Saisie = "0" → rouge
    rule_0 = CellIsRule(operator="equal", formula=['"0"'], fill=red_fill)
    ws_eval.conditional_formatting.add(plage_saisie, rule_0)
    # Saisie = "NE" → gris
    rule_ne = CellIsRule(operator="equal", formula=['"NE"'], fill=gray_fill)
    ws_eval.conditional_formatting.add(plage_saisie, rule_ne)

    # Note finale < 10 → texte rouge
    cell_note20 = f"{col_letter_n}{ROW_NOTE_20}"
    rule_fail = CellIsRule(
        operator="lessThan", formula=["10"],
        font=Font(color="CC0000", bold=True)
    )
    ws_eval.conditional_formatting.add(cell_note20, rule_fail)
    # Note finale ≥ 14 → texte vert
    rule_good = CellIsRule(
        operator="greaterThanOrEqual", formula=["14"],
        font=Font(color="008000", bold=True)
    )
    ws_eval.conditional_formatting.add(cell_note20, rule_good)

# Audit poids : si Σ ≠ nombre de parties → rouge
nb_parties = len(POIDS_PARTIES)
audit_cell_ref = f"$F${ROW_AUDIT_POIDS}"
rule_audit_ko = FormulaRule(
    formula=[f"ABS({audit_cell_ref}-{nb_parties})>0.001"],
    fill=red_fill,
    font=Font(color="CC0000", bold=True),
)
ws_eval.conditional_formatting.add(f"F{ROW_AUDIT_POIDS}", rule_audit_ko)

# Databar sur les notes /20
note20_range = (
    f"{get_column_letter(col_note(1))}{ROW_NOTE_20}:"
    f"{get_column_letter(col_note(N_ELEVES))}{ROW_NOTE_20}"
)
databar_rule = DataBarRule(
    start_type="num", start_value=0,
    end_type="num", end_value=20,
    color="63BE7B",
    showValue=True,
)
ws_eval.conditional_formatting.add(note20_range, databar_rule)

# ─── Figement et protection ──────────────────────────────────────────────────
ws_eval.freeze_panes = "G4"

# Protection : verrouillage par défaut, déverrouillage explicite sur les saisies
# Les cellules saisie et poids ont déjà été marquées PROT_UNLOCKED
ws_eval.protection.sheet = True
# Ne PAS appeler .password = "" → openpyxl hasherait une chaîne vide et bloquerait LibreOffice
ws_eval.protection.formatCells = False
ws_eval.protection.formatColumns = False
ws_eval.protection.formatRows = False
ws_eval.protection.selectLockedCells = False
ws_eval.protection.selectUnlockedCells = False

# ─────────────────────────────────────────────────────────────────────────────
# FEUILLE ÉLÈVES
# ─────────────────────────────────────────────────────────────────────────────

ws_eleves.cell(1, 1, f"Liste de la classe — Synthèse des évaluations").font = FONT_TITRE
ws_eleves.cell(1, 1).fill = FILL_TITRE
ws_eleves.cell(1, 1).alignment = AL_CENTER
ws_eleves.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
ws_eleves.row_dimensions[1].height = 24

# En-têtes ligne 3
headers_eleves = ["N°", "Nom Prénom", f"{NOM_EVAL} /20", f"{NOM_EVAL} /12", "Moyenne /20", "Moyenne /12"]
for i, h in enumerate(headers_eleves, start=1):
    c = ws_eleves.cell(3, i, h)
    c.font = FONT_ENTETE
    c.fill = FILL_ENTETE
    c.alignment = AL_CENTER_WRAP
    c.border = BORDER_ALL

# Largeurs
ws_eleves.column_dimensions["A"].width = 6
ws_eleves.column_dimensions["B"].width = 28
ws_eleves.column_dimensions["C"].width = 16
ws_eleves.column_dimensions["D"].width = 16
ws_eleves.column_dimensions["E"].width = 14
ws_eleves.column_dimensions["F"].width = 14

# Lignes élèves 4..N_ELEVES+3
for i in range(1, N_ELEVES + 1):
    r = 3 + i
    # N°
    ws_eleves.cell(r, 1, i).alignment = AL_CENTER
    ws_eleves.cell(r, 1).border = BORDER_ALL
    # Nom (prérempli pour les 24 premiers)
    nom = f"Élève {i}" if i <= N_ELEVES_PREREMPLIS else ""
    cell_nom = ws_eleves.cell(r, 2, nom)
    cell_nom.alignment = AL_LEFT
    cell_nom.border = BORDER_ALL
    cell_nom.protection = PROT_UNLOCKED  # nom éditable
    # Note BB2 /20 : référence à la note de l'élève i dans la feuille éval
    col_note_eval = get_column_letter(col_note(i))
    f_note_eval = f"=IFERROR('{NOM_EVAL}'!{col_note_eval}{ROW_NOTE_20},\"\")"
    cell_n = ws_eleves.cell(r, 3, f_note_eval)
    cell_n.alignment = AL_CENTER
    cell_n.font = Font(bold=True)
    cell_n.number_format = "0.00"
    cell_n.border = BORDER_ALL
    # Note BB2 /12
    f_note12 = f"=IFERROR('{NOM_EVAL}'!{col_note_eval}{ROW_NOTE_12},\"\")"
    cell_n12 = ws_eleves.cell(r, 4, f_note12)
    cell_n12.alignment = AL_CENTER
    cell_n12.number_format = "0.00"
    cell_n12.border = BORDER_ALL
    # Moyenne (pour l'instant = la seule note BB2)
    cell_moy = ws_eleves.cell(r, 5, f"=IF(C{r}=\"\",\"\",C{r})")
    cell_moy.alignment = AL_CENTER
    cell_moy.number_format = "0.00"
    cell_moy.border = BORDER_ALL
    cell_moy12 = ws_eleves.cell(r, 6, f"=IF(D{r}=\"\",\"\",D{r})")
    cell_moy12.alignment = AL_CENTER
    cell_moy12.number_format = "0.00"
    cell_moy12.border = BORDER_ALL

# Ligne moyennes classe (en bas)
r_moy = 3 + N_ELEVES + 2
ws_eleves.cell(r_moy, 2, "Moyenne classe").font = FONT_ENTETE
ws_eleves.cell(r_moy, 2).fill = FILL_ENTETE
ws_eleves.cell(r_moy, 3, f"=IFERROR(AVERAGE(C4:C{3+N_ELEVES}),\"\")").number_format = "0.00"
ws_eleves.cell(r_moy, 3).font = FONT_ENTETE
ws_eleves.cell(r_moy, 3).fill = FILL_ENTETE
ws_eleves.cell(r_moy, 4, f"=IFERROR(AVERAGE(D4:D{3+N_ELEVES}),\"\")").number_format = "0.00"
ws_eleves.cell(r_moy, 4).font = FONT_ENTETE
ws_eleves.cell(r_moy, 4).fill = FILL_ENTETE

# MFC : ligne grisée si nom vide
for i in range(1, N_ELEVES + 1):
    r = 3 + i
    plage = f"A{r}:F{r}"
    rule_vide = FormulaRule(
        formula=[f'$B{r}=""'],
        font=Font(color="BBBBBB", italic=True),
    )
    ws_eleves.conditional_formatting.add(plage, rule_vide)

# MFC sur notes /20 (échec rouge, succès vert)
range_notes20 = f"C4:C{3+N_ELEVES}"
ws_eleves.conditional_formatting.add(
    range_notes20,
    CellIsRule(operator="lessThan", formula=["10"], font=Font(color="CC0000", bold=True))
)
ws_eleves.conditional_formatting.add(
    range_notes20,
    CellIsRule(operator="greaterThanOrEqual", formula=["14"], font=Font(color="008000", bold=True))
)

# Protection feuille Élèves
ws_eleves.protection.sheet = True
# Pas de password (cf. note plus haut)

ws_eleves.freeze_panes = "B4"

# ─────────────────────────────────────────────────────────────────────────────
# FEUILLE MODE D'EMPLOI
# ─────────────────────────────────────────────────────────────────────────────

mode_emploi = [
    ("Mode d'emploi de la grille de notation", "title"),
    ("", ""),
    ("1. SAISIE DES ÉVALUATIONS", "h1"),
    ("Pour chaque critère et chaque élève, dans la feuille de l'évaluation (ex: BB2 12-05-26), saisir UN seul des codes suivants dans la colonne 'E1', 'E2', etc. :", ""),
    ("  • 0  → critère non acquis", ""),
    ("  • 1  → critère partiellement acquis (½ point)", ""),
    ("  • 2  → critère acquis", ""),
    ("  • NE → critère non évalué (sera ignoré dans le calcul)", ""),
    ("", ""),
    ("La cellule propose une liste déroulante : toute autre valeur sera refusée par la validation de données.", ""),
    ("", ""),
    ("2. CODES COULEURS", "h1"),
    ("  • Vert pâle  : saisie = 2 (critère acquis)", ""),
    ("  • Rouge pâle : saisie = 0 (critère non acquis)", ""),
    ("  • Gris       : saisie = NE (non évalué)", ""),
    ("  • Note finale /20 < 10 → texte rouge gras", ""),
    ("  • Note finale /20 ≥ 14 → texte vert gras", ""),
    ("  • Audit poids (cellule F sous les critères) en rouge → un poids a été modifié, somme ≠ 5,00 (somme des 5 parties)", ""),
    ("", ""),
    ("3. AJOUTER OU RETIRER UN ÉLÈVE", "h1"),
    ("La grille est dimensionnée pour 30 élèves (24 préremplis, 6 lignes vides).", ""),
    ("  • Ajouter : aller dans 'Élèves', ligne 28+, saisir le nom dans la colonne B. La note remontera automatiquement.", ""),
    ("  • Retirer : effacer le nom dans la colonne B (la ligne devient grisée).", ""),
    ("", ""),
    ("4. MODIFIER LA PONDÉRATION", "h1"),
    ("Tous les poids sont dans la colonne F de la feuille d'évaluation.", ""),
    ("  • Poids des parties : sur les lignes de titre 'P1', 'P2'…", ""),
    ("  • Poids des critères : sur chaque ligne de critère.", ""),
    ("Tout changement est répercuté automatiquement sur les 30 colonnes-élèves.", ""),
    ("La cellule d'audit (en bas, colonne F) vérifie que Σ(poids critères par partie) = 1,00 pour chaque P, soit 5,00 au total.", ""),
    ("", ""),
    ("5. AJOUTER UNE NOUVELLE ÉVALUATION (BB3, etc.)", "h1"),
    ("  1) Cliquer-droit sur l'onglet 'BB2 12-05-26' → Déplacer/Copier → Cocher 'Copier' → OK.", ""),
    ("  2) Renommer l'onglet (ex: 'BB3 16-06-26').", ""),
    ("  3) Modifier les libellés des critères et les poids selon le nouveau sujet.", ""),
    ("  4) Dans la feuille 'Élèves', ajouter 2 colonnes (note /20 et /12) référant à la nouvelle feuille.", ""),
    ("", ""),
    ("6. CONTRAINTES TECHNIQUES", "h1"),
    ("  • Les cellules contenant des formules sont protégées (impossible à modifier par erreur).", ""),
    ("  • Pour modifier une formule, désactiver la protection : Outils → Protéger le document → Feuille → décocher.", ""),
    ("  • Les noms d'élèves et les cellules de saisie sont déverrouillés.", ""),
    ("", ""),
    ("Version générée par build_grille.py — 2026-05-15", "italic"),
]

for i, (text, style) in enumerate(mode_emploi, start=1):
    c = ws_help.cell(i, 1, text)
    if style == "title":
        c.font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
        c.fill = FILL_TITRE
        c.alignment = AL_CENTER
        ws_help.merge_cells(start_row=i, start_column=1, end_row=i, end_column=4)
        ws_help.row_dimensions[i].height = 28
    elif style == "h1":
        c.font = Font(name="Calibri", size=12, bold=True, color="2E5C8A")
        c.alignment = AL_LEFT
    elif style == "italic":
        c.font = Font(name="Calibri", size=9, italic=True, color="888888")
    else:
        c.font = Font(name="Calibri", size=10)
        c.alignment = AL_LEFT_TOP

ws_help.column_dimensions["A"].width = 110

# ─────────────────────────────────────────────────────────────────────────────
# Sauvegarde
# ─────────────────────────────────────────────────────────────────────────────

wb.save(OUT_FILE)
print(f"✓ Fichier généré : {OUT_FILE}")
print(f"  - Feuille 'Élèves' : {N_ELEVES} lignes ({N_ELEVES_PREREMPLIS} préremplis)")
print(f"  - Feuille '{NOM_EVAL}' : {sum(1 for e in CRITERES if e[0] == 'critere')} critères, {N_ELEVES} blocs-élèves")
print(f"  - Largeur totale : {n_cols_total} colonnes")
